from openpyxl.styles import Font, colors, fills
from openpyxl.workbook import Workbook
from openpyxl.chart import DoughnutChart, Reference


class TestCase:
    def __init__(self, tca, result, exitcode):
        self._testCaseName = tca
        self._exitCode = exitcode
        self._result = result


class Module:
    def __init__(self, name):
        self._name = name

        """"List of Test Cases"""
        self._listTestCases = []

    def append_test_case(self, tca):
        self._listTestCases.append(tca)


class ReportData:
    def __init__(self):
        """List of Modules"""
        self._listModules = []
        self.totalTests = ''
        self.skippedTests = ''
        self.totalFailures = ''
        self.nrTotalTest = 0
        self.nrTotalFailures = 0
        self.nrTotalSkipped = 0
        self.nrTotalPass = 0
        self.percentagePass = 0
        self.percentageConf = 0
        self.percentageFail = 0

    def append_module(self, module):
        self._listModules.append(module)

    def append_total_tests(self, total_str):
        self.totalTests += total_str

    def append_skipped_tests(self, skipped_str):
        self.skippedTests += skipped_str

    def append_total_failures(self, failures_str):
        self.totalFailures += failures_str


class Generator:
    pass_str = 'PASS'
    fail_str = 'FAIL'
    conf_str = 'CONF'
    total_tests = 'Total Tests'
    skipped_test = 'Total Skipped'
    total_failures = 'Total Failures'

    @staticmethod
    def file_parser_ltp(report_file, report):
        with open(report_file, 'r') as file:
            for line in file:
                if Generator.pass_str in line or Generator.fail_str in line or Generator.conf_str in line:
                    words = line.split()
                    if len(words) >= 3:
                        test_case = TestCase(words[0], words[1], words[2])
                        module = Module(words[0])
                        module.append_test_case(test_case)
                        report.append_module(module)
                if Generator.total_tests in line:
                    report.append_total_tests(line)
                    words = line.split()
                    report.nrTotalTest += int(words[2])

                if Generator.skipped_test in line:
                    report.append_skipped_tests(line)
                    words = line.split()
                    report.nrTotalSkipped += int(words[3])

                if Generator.total_failures in line:
                    report.append_total_failures(line)
                    words = line.split()
                    report.nrTotalFailures += int(words[2])
            report.nrTotalPass = report.nrTotalTest - report.nrTotalFailures - report.nrTotalSkipped

        report.percentagePass = round(report.nrTotalPass * 100 / report.nrTotalTest, 2)
        report.percentageConf = round(report.nrTotalSkipped * 100 / report.nrTotalTest, 2)
        report.percentageFail = round(report.nrTotalFailures * 100 / report.nrTotalTest, 2)

    @staticmethod
    def list_test_cases(report):
        for module in report._listModules:
            print('Module name: ', module._name)
            for tca in module._listTestCases:
                print('TestCase: ', 'name: ', tca._testCaseName, ' result:', tca._result, ' exit code:', tca._exitCode)

    @staticmethod
    def append_data_into_cells(worksheet, report):
        current_row = 5
        current_column = 1
        my_red = colors.Color(colors.RED)
        my_fill_red = fills.PatternFill(patternType='solid', fgColor=my_red)
        my_green = colors.Color(colors.GREEN)
        my_fill_green = fills.PatternFill(patternType='solid', fgColor=my_green)
        my_pink = colors.Color(rgb='FF9999')
        my_fill_pink = fills.PatternFill(patternType='solid', fgColor=my_pink)

        for module in report._listModules:
            worksheet.cell(row=current_row, column=current_column).value = module._name
            for tca in module._listTestCases:
                current_column += 1
                worksheet.cell(row=current_row, column=current_column).value = tca._testCaseName
                current_column += 1
                worksheet.cell(row=current_row, column=current_column).value = tca._result
                if Generator.pass_str in tca._result:
                    worksheet.cell(row=current_row, column=current_column).fill = my_fill_green
                elif Generator.fail_str in tca._result:
                    worksheet.cell(row=current_row, column=current_column).fill = my_fill_red
                else:
                    worksheet.cell(row=current_row, column=current_column).fill = my_fill_pink
                current_column += 1
                worksheet.cell(row=current_row, column=current_column).value = tca._exitCode
                current_column = 1
            current_row += 1

        current_column = 1
        current_row += 1
        worksheet.cell(row=current_row, column=current_column).value = 'Summary:'
        current_row += 1
        worksheet.cell(row=current_row, column=current_column).value = Generator.total_tests
        current_column += 1
        worksheet.cell(row=current_row, column=current_column).value = report.nrTotalTest
        current_column = 1
        current_row += 1
        worksheet.cell(row=current_row, column=current_column).value = Generator.skipped_test
        current_column += 1
        worksheet.cell(row=current_row, column=current_column).value = report.nrTotalSkipped
        current_row += 1
        current_column = 1
        worksheet.cell(row=current_row, column=current_column).value = Generator.total_failures
        current_column += 1
        worksheet.cell(row=current_row, column=current_column).value = report.nrTotalFailures
        current_column = 1
        current_row += 1
        worksheet.cell(row=current_row, column=current_column).value = 'Percentage Pass'
        worksheet.cell(row=current_row, column=current_column).fill = my_fill_green
        current_column += 1
        worksheet.cell(row=current_row, column=current_column).value = report.percentagePass
        worksheet.cell(row=current_row, column=current_column).fill = my_fill_green
        current_column = 1
        current_row += 1
        worksheet.cell(row=current_row, column=current_column).value = 'Percentage Fail'
        worksheet.cell(row=current_row, column=current_column).fill = my_fill_red
        current_column += 1
        worksheet.cell(row=current_row, column=current_column).value = report.percentageFail
        worksheet.cell(row=current_row, column=current_column).fill = my_fill_red
        current_column = 1
        current_row += 1
        worksheet.cell(row=current_row, column=current_column).value = 'Percentage Conf'
        worksheet.cell(row=current_row, column=current_column).fill = my_fill_pink
        current_column += 1
        worksheet.cell(row=current_row, column=current_column).value = report.percentageConf
        worksheet.cell(row=current_row, column=current_column).fill = my_fill_pink

        current_row -= 2
        current_column = 1
        # at the end create a chart
        chart = DoughnutChart()
        labels = Reference(worksheet, min_col=current_column, min_row=current_row, max_row=current_row + 2)
        data = Reference(worksheet, min_col=current_column + 1, min_row=current_row, max_row=current_row + 2)
        chart.add_data(data)
        chart.set_categories(labels)
        chart.title = "LTP test results"
        # Change bar filling and line color

        # serie1= chart.series[0];
        # serie1.graphicalProperties.solidFill = "7E3F00"
        # serie2 = chart.series[1];

        chart.style = 2
        worksheet.add_chart(chart, "F3")

    @staticmethod
    def format_excel_sheet(report, ltp_file):
        # Create a workbook
        workbook = Workbook()
        sheet = workbook.active

        bold_font = Font(bold=True, color=colors.DARKYELLOW, size=20)

        # set the width of the column
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['C'].width = 10

        sheet['A1'].font = bold_font

        sheet.merge_cells('A1:D1')

        sheet['A1'] = 'LTP Test report'

        my_yellow = colors.Color('e6e600')
        my_fill_yellow = fills.PatternFill(patternType='solid', fgColor=my_yellow)

        sheet['A3'] = 'Module'
        sheet.cell(row=3, column=1).fill = my_fill_yellow
        sheet['B3'] = 'Test Case'
        sheet.cell(row=3, column=2).fill = my_fill_yellow
        sheet['C3'] = 'Result'
        sheet.cell(row=3, column=3).fill = my_fill_yellow
        sheet['D3'] = 'Exit Code'
        sheet.cell(row=3, column=4).fill = my_fill_yellow

        Generator.append_data_into_cells(sheet, report)
        filename_ltp = ltp_file.split('\\')
        filename_ltp_no_ext = (filename_ltp[-1].split('.'))[0]
        output_file = 'l4b-software___testReport' + '___' + filename_ltp_no_ext + '.xlsx'

        try:
            workbook.save(filename=output_file)
        except PermissionError as e:
            print("\n\n\n Excel file is open. Please close the excel file !!!")